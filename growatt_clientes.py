"""
Importador mensual de datos Growatt para la base de clientes.

Uso básico:
    export GROWATT_API_TOKEN="tu_token"
    python3 growatt_clientes.py --mes actual

Alternativa sin token:
    export GROWATT_SERVER_COOKIE="JSESSIONID=...; SERVERID=..."
    python3 growatt_clientes.py --mes actual

Opciones útiles:
    python3 growatt_clientes.py --mes 2026-03 --salida data/growatt_clientes_2026-03.xlsx
    python3 growatt_clientes.py --mes 2026-03 --actualizar-datos
    python3 growatt_clientes.py --mes 2026-03 --actualizar-datos --solo-actualizar-datos
    python3 growatt_clientes.py --mes 2026-03 --cargar-auxiliar
    python3 growatt_clientes.py --cargar-todos-auxiliares

El script:
    1. Se conecta a Growatt
    2. Obtiene la lista de plantas
    3. Filtra solo las plantas incluidas en data/growatt_allowlist.csv
       (columna obligatoria: plant_id; columnas extra opcionales)
    4. Consulta la generación del mes por planta
    5. Genera un Excel auxiliar con el detalle
    6. Opcionalmente actualiza la hoja "datos" de data/datos_clientes.xlsx
       cargando o actualizando la columna "generacion_kwh" por cliente y mes
"""

from __future__ import annotations

import argparse
import calendar
import html
import os
import sys
import time
import csv
import re
import smtplib
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from email.message import EmailMessage
from pathlib import Path

import growattServer
import openpyxl
import pandas as pd
import requests
from openpyxl.styles import Alignment, Font, PatternFill


DEFAULT_XLSX_DATOS = Path("data/datos_clientes.xlsx")
DEFAULT_PAUSA_SEG = 1.5
DEFAULT_REINTENTOS = 3
DEFAULT_API_TOKEN = "bqp7dd06xcjfs06j51402m004wxpy5jv"
DEFAULT_ALLOWLIST_PATH = Path("data/growatt_allowlist.csv")
DEFAULT_SERVER_URL = "https://server.growatt.com/"
DEFAULT_OPENAPI_BASE_URL = "https://openapi.growatt.com/v1"
DEFAULT_ALERT_EMAIL_TO = "nmachin@voltia.com.uy"
DEFAULT_SMTP_HOST = "mail.voltia.com.uy"
DEFAULT_SMTP_PORT = 465
DEFAULT_SMTP_USER = "reportes@voltia.com.uy"
DEFAULT_SMTP_PASSWORD = "Voltia123"
DEFAULT_SMTP_FROM = "reportes@voltia.com.uy"


@dataclass
class Config:
    api_token: str
    server_cookie: str
    usuario: str
    password: str
    mes: str
    pausa_seg: float
    reintentos: int
    salida: Path | None
    actualizar_datos: bool
    solo_actualizar_datos: bool
    xlsx_datos: Path
    allowlist_path: Path
    listar_plantas: bool
    diagnostico_cookie: bool
    cargar_todos_auxiliares: bool
    alert_email_to: str


def parse_args() -> Config:
    parser = argparse.ArgumentParser(
        description="Importa generación mensual desde Growatt y la deja lista para el reporte."
    )
    parser.add_argument(
        "--mes",
        default=os.environ.get("GROWATT_MONTH", "anterior"),
        help='Mes a consultar en formato "YYYY-MM", "actual" o "anterior".',
    )
    parser.add_argument(
        "--pausa",
        type=float,
        default=float(os.environ.get("GROWATT_PAUSE_SECONDS", DEFAULT_PAUSA_SEG)),
        help="Pausa entre requests a Growatt en segundos.",
    )
    parser.add_argument(
        "--reintentos",
        type=int,
        default=int(os.environ.get("GROWATT_RETRIES", DEFAULT_REINTENTOS)),
        help="Cantidad de reintentos por planta ante errores transitorios.",
    )
    parser.add_argument(
        "--salida",
        type=Path,
        default=None,
        help="Ruta del Excel auxiliar a generar.",
    )
    parser.add_argument(
        "--actualizar-datos",
        action="store_true",
        help='Actualiza la hoja "datos" del archivo base con generacion_kwh.',
    )
    parser.add_argument(
        "--solo-actualizar-datos",
        action="store_true",
        help=(
            "No consulta Growatt. Lee el Excel auxiliar existente (según --salida o el nombre "
            "por defecto del mes) y actualiza la hoja \"datos\"."
        ),
    )
    parser.add_argument(
        "--cargar-auxiliar",
        action="store_true",
        help=(
            "Atajo de --actualizar-datos --solo-actualizar-datos. "
            "Carga en la hoja \"datos\" el Excel auxiliar ya generado."
        ),
    )
    parser.add_argument(
        "--cargar-todos-auxiliares",
        action="store_true",
        help=(
            "No consulta Growatt. Carga en la hoja \"datos\" todos los Excel auxiliares "
            "generados en data/."
        ),
    )
    parser.add_argument(
        "--xlsx-datos",
        type=Path,
        default=DEFAULT_XLSX_DATOS,
        help='Ruta al archivo Excel que contiene la hoja "datos".',
    )
    parser.add_argument(
        "--allowlist",
        type=Path,
        default=DEFAULT_ALLOWLIST_PATH,
        help="CSV con plantas autorizadas. La unica columna obligatoria es plant_id.",
    )
    parser.add_argument(
        "--listar-plantas",
        action="store_true",
        help="Lista las plantas visibles por el token y termina sin exportar.",
    )
    parser.add_argument(
        "--diagnostico-cookie",
        action="store_true",
        help="Muestra la respuesta cruda de server.growatt.com al usar la cookie web.",
    )
    parser.add_argument(
        "--alert-email-to",
        default=os.environ.get("GROWATT_ALERT_EMAIL_TO", DEFAULT_ALERT_EMAIL_TO),
        help="Casilla a la que se enviarán alertas automáticas por exportación anómala.",
    )

    args = parser.parse_args()

    actualizar_datos = args.actualizar_datos or args.cargar_auxiliar or args.cargar_todos_auxiliares
    solo_actualizar_datos = args.solo_actualizar_datos or args.cargar_auxiliar or args.cargar_todos_auxiliares

    api_token = os.environ.get("GROWATT_API_TOKEN", DEFAULT_API_TOKEN).strip()
    server_cookie = os.environ.get("GROWATT_SERVER_COOKIE", "").strip()
    usuario = os.environ.get("GROWATT_USER", "").strip()
    password = os.environ.get("GROWATT_PASSWORD", "").strip()

    if not api_token and not server_cookie and (not usuario or not password):
        raise SystemExit(
            "Faltan credenciales. Definí GROWATT_API_TOKEN, GROWATT_SERVER_COOKIE o bien GROWATT_USER y GROWATT_PASSWORD antes de ejecutar."
        )

    return Config(
        api_token=api_token,
        server_cookie=server_cookie,
        usuario=usuario,
        password=password,
        mes=args.mes,
        pausa_seg=args.pausa,
        reintentos=args.reintentos,
        salida=args.salida,
        actualizar_datos=actualizar_datos,
        solo_actualizar_datos=solo_actualizar_datos,
        xlsx_datos=args.xlsx_datos,
        allowlist_path=args.allowlist,
        listar_plantas=args.listar_plantas,
        diagnostico_cookie=args.diagnostico_cookie,
        cargar_todos_auxiliares=args.cargar_todos_auxiliares,
        alert_email_to=args.alert_email_to.strip(),
    )


def obtener_fecha_inicio(mes: str) -> date:
    if mes.lower() == "actual":
        hoy = date.today()
        return date(hoy.year, hoy.month, 1)
    if mes.lower() == "anterior":
        hoy = date.today()
        primer_dia_mes_actual = date(hoy.year, hoy.month, 1)
        ultimo_dia_mes_anterior = primer_dia_mes_actual - timedelta(days=1)
        return date(ultimo_dia_mes_anterior.year, ultimo_dia_mes_anterior.month, 1)
    return datetime.strptime(mes, "%Y-%m").date()


def obtener_nombre_salida(fecha: date) -> Path:
    return Path(f"data/growatt_clientes_{fecha.strftime('%Y-%m')}.xlsx")


def a_float(valor):
    if valor in (None, "", "N/D"):
        return None
    try:
        return float(valor)
    except (TypeError, ValueError):
        return None


def formatear_progreso(indice: int, total: int) -> str:
    if total <= 0:
        return "[0/0 | 0%]"
    porcentaje = round((indice / total) * 100)
    return f"[{indice}/{total} | {porcentaje}%]"


def formatear_metrica(valor) -> str:
    return "N/D" if valor is None else str(valor)


def construir_html_alerta_exportacion_baja(alertas: list[dict], meses: list[str]) -> str:
    filas_html = []
    for alerta in alertas:
        filas_html.append(
            "<tr>"
            f"<td>{html.escape(str(alerta['cliente']))}</td>"
            f"<td>{html.escape(str(alerta['id_planta']))}</td>"
            f"<td>{html.escape(str(alerta['mes']))}</td>"
            f"<td>{html.escape(formatear_metrica(alerta['generacion_kwh']))}</td>"
            f"<td>{html.escape(formatear_metrica(alerta['consumo_kwh']))}</td>"
            f"<td>{html.escape(formatear_metrica(alerta['exportacion_kwh']))}</td>"
            "</tr>"
        )

    chips_html = "".join(
        f'<span style="display:inline-block;background:#e9f3ff;color:#0f4c81;'
        f'padding:6px 10px;border-radius:999px;font-size:12px;font-weight:600;'
        f'margin:0 8px 8px 0;">{html.escape(mes)}</span>'
        for mes in meses
    )

    return f"""
    <html>
      <body style="margin:0;padding:24px;background:#f4f7fb;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;color:#1f2937;">
        <div style="max-width:900px;margin:0 auto;background:#ffffff;border:1px solid #dbe4f0;border-radius:18px;overflow:hidden;">
          <div style="background:linear-gradient(135deg,#0f4c81 0%,#2e75b6 100%);padding:28px 32px;color:#ffffff;">
            <div style="font-size:13px;letter-spacing:.08em;text-transform:uppercase;opacity:.85;">Voltia · Monitoreo Growatt</div>
            <h1 style="margin:10px 0 8px 0;font-size:28px;line-height:1.2;">Posible exportación cero configurada</h1>
            <p style="margin:0;font-size:15px;line-height:1.6;opacity:.92;">
              Se detectaron clientes con generación mayor a 100 kWh y exportación menor a 5 kWh.
            </p>
          </div>

          <div style="padding:28px 32px 12px 32px;">
            <div style="margin-bottom:18px;">{chips_html}</div>

            <div style="background:#fff7ed;border:1px solid #fed7aa;border-radius:14px;padding:18px 20px;margin-bottom:22px;">
              <div style="font-weight:700;color:#9a3412;margin-bottom:8px;">Criterio de alerta</div>
              <div style="font-size:14px;line-height:1.7;color:#7c2d12;">
                <div>Generación mayor a <strong>100 kWh</strong></div>
                <div>Exportación menor a <strong>5 kWh</strong></div>
                <div>Exportación con dato válido</div>
              </div>
            </div>

            <div style="font-size:16px;font-weight:700;color:#111827;margin-bottom:12px;">
              Clientes detectados ({len(alertas)})
            </div>

            <div style="overflow-x:auto;border:1px solid #dbe4f0;border-radius:14px;">
              <table style="width:100%;border-collapse:collapse;font-size:14px;">
                <thead>
                  <tr style="background:#edf4fb;color:#0f4c81;text-align:left;">
                    <th style="padding:12px 14px;border-bottom:1px solid #dbe4f0;">Cliente</th>
                    <th style="padding:12px 14px;border-bottom:1px solid #dbe4f0;">Planta</th>
                    <th style="padding:12px 14px;border-bottom:1px solid #dbe4f0;">Mes</th>
                    <th style="padding:12px 14px;border-bottom:1px solid #dbe4f0;">Generación</th>
                    <th style="padding:12px 14px;border-bottom:1px solid #dbe4f0;">Consumo</th>
                    <th style="padding:12px 14px;border-bottom:1px solid #dbe4f0;">Exportación</th>
                  </tr>
                </thead>
                <tbody>
                  {"".join(filas_html)}
                </tbody>
              </table>
            </div>

            <p style="margin:18px 0 0 0;font-size:13px;line-height:1.6;color:#6b7280;">
              Este aviso se genera automáticamente al actualizar <strong>datos_clientes.xlsx</strong>.
            </p>
          </div>
        </div>
      </body>
    </html>
    """


def detectar_alertas_exportacion_baja(filas: list[dict]) -> list[dict]:
    alertas = []
    for fila in filas:
        generacion = a_float(fila.get("generacion_kwh"))
        exportacion = a_float(fila.get("exportacion_kwh"))
        if generacion is None or exportacion is None:
            continue
        if generacion > 100 and exportacion < 5:
            alertas.append(
                {
                    "cliente": str(fila.get("cliente") or "").strip(),
                    "id_planta": fila.get("id_planta"),
                    "mes": str(fila.get("mes") or "").strip(),
                    "generacion_kwh": generacion,
                    "consumo_kwh": a_float(fila.get("consumo_kwh")),
                    "exportacion_kwh": exportacion,
                }
            )
    return alertas


def enviar_alerta_exportacion_baja(config: Config, alertas: list[dict]) -> None:
    if not alertas:
        return

    smtp_host = os.environ.get("SMTP_HOST", DEFAULT_SMTP_HOST).strip()
    smtp_port = int(os.environ.get("SMTP_PORT", str(DEFAULT_SMTP_PORT)))
    smtp_user = os.environ.get("SMTP_USER", DEFAULT_SMTP_USER).strip()
    smtp_password = os.environ.get("SMTP_PASSWORD", DEFAULT_SMTP_PASSWORD).strip()
    smtp_from = os.environ.get("SMTP_FROM", DEFAULT_SMTP_FROM).strip()
    smtp_use_ssl = os.environ.get("SMTP_USE_SSL", "1").strip() != "0"
    smtp_use_tls = os.environ.get("SMTP_USE_TLS", "0").strip() != "0"

    meses = sorted({alerta["mes"] for alerta in alertas if alerta.get("mes")})
    asunto = (
        f"Alerta Growatt: posible exportación cero configurada ({', '.join(meses)})"
        if meses else
        "Alerta Growatt: posible exportación cero configurada"
    )

    lineas = [
        "Se detectaron clientes con posible exportación cero configurada.",
        "",
        "Criterio:",
        "- generacion_kwh > 100",
        "- exportacion_kwh < 5",
        "- exportacion_kwh con dato válido",
        "",
        "Clientes detectados:",
    ]

    for alerta in alertas:
        lineas.append(
            "- "
            f"{alerta['cliente']} | planta={alerta['id_planta']} | mes={alerta['mes']} | "
            f"gen={alerta['generacion_kwh']} | cons={formatear_metrica(alerta['consumo_kwh'])} | "
            f"exp={alerta['exportacion_kwh']}"
        )

    mensaje = EmailMessage()
    mensaje["Subject"] = asunto
    mensaje["From"] = smtp_from
    mensaje["To"] = config.alert_email_to
    mensaje.set_content("\n".join(lineas))
    mensaje.add_alternative(
        construir_html_alerta_exportacion_baja(alertas, meses),
        subtype="html",
    )

    try:
        cliente_smtp = smtplib.SMTP_SSL if smtp_use_ssl else smtplib.SMTP
        with cliente_smtp(smtp_host, smtp_port, timeout=30) as servidor:
            if not smtp_use_ssl:
                servidor.ehlo()
                if smtp_use_tls:
                    servidor.starttls()
                    servidor.ehlo()
            servidor.login(smtp_user, smtp_password)
            servidor.send_message(mensaje)
        print(f"Alerta enviada por email a {config.alert_email_to}: {len(alertas)} cliente(s).")
    except Exception as exc:  # noqa: BLE001
        print(f"AVISO: no se pudo enviar la alerta por email: {exc}")


def enviar_correo_prueba_alerta(config: Config) -> None:
    alerta_prueba = [{
        "cliente": "PRUEBA ALERTA",
        "id_planta": "TEST",
        "mes": date.today().strftime("%Y-%m"),
        "generacion_kwh": 150.0,
        "consumo_kwh": 120.0,
        "exportacion_kwh": 0.0,
    }]
    enviar_alerta_exportacion_baja(config, alerta_prueba)


def ultimo_dia_mes(fecha: date) -> date:
    return date(fecha.year, fecha.month, calendar.monthrange(fecha.year, fecha.month)[1])


def resumir_historial_diario(datas: list[dict], campo: str) -> float | None:
    maximos_por_dia = {}
    for item in datas:
        marca_tiempo = str(item.get("time") or item.get("timeText") or "").strip()
        if not marca_tiempo:
            continue
        dia = marca_tiempo[:10]
        valor = a_float(item.get(campo))
        if valor is None:
            continue
        anterior = maximos_por_dia.get(dia)
        if anterior is None or valor > anterior:
            maximos_por_dia[dia] = valor

    if not maximos_por_dia:
        return None

    return round(sum(maximos_por_dia.values()), 2)


def openapi_get_data(
    api: growattServer.OpenApiV1,
    endpoint: str,
    params: dict,
    descripcion: str,
    reintentos: int = 5,
) -> dict:
    ultimo_error = None
    for intento in range(1, reintentos + 1):
        try:
            response = api.session.get(api.get_url(endpoint), params=params)
            response.raise_for_status()
            try:
                payload = response.json()
            except ValueError as exc:
                texto = response.text[:500].strip()
                raise RuntimeError(
                    f"{descripcion}: respuesta no JSON (HTTP {response.status_code}). "
                    f"Body: {texto}"
                ) from exc

            try:
                return api.process_response(payload, descripcion)
            except Exception as exc:
                error_code = payload.get("error_code")
                error_msg = payload.get("error_msg")
                raise RuntimeError(
                    f"{descripcion}: error_code={error_code} error_msg={error_msg} "
                    f"endpoint={endpoint} params={params}"
                ) from exc
        except Exception as exc:  # noqa: BLE001
            ultimo_error = exc
            if intento < reintentos:
                time.sleep(min(1.5 * intento, 5.0))
    raise ultimo_error if ultimo_error is not None else RuntimeError(descripcion)


def obtener_generacion_mensual_directa(
    api_token: str,
    plant_id: int,
    fecha: date,
    fecha_fin: date,
    reintentos: int = 5,
) -> float | None:
    ultimo_error = None
    for intento in range(1, reintentos + 1):
        try:
            response = requests.get(
                f"{DEFAULT_OPENAPI_BASE_URL}/plant/energy",
                headers={"Token": api_token},
                params={
                    "plant_id": plant_id,
                    "start_date": fecha.strftime("%Y-%m-%d"),
                    "end_date": fecha_fin.strftime("%Y-%m-%d"),
                    "time_unit": "month",
                },
                timeout=30,
            )
            response.raise_for_status()
            payload = response.json()
            if payload.get("error_code") != 0:
                raise RuntimeError(
                    f"Growatt devolvió error_code={payload.get('error_code')} error_msg={payload.get('error_msg')}"
                )
            energys = payload.get("data", {}).get("energys", [])
            energia_kwh = a_float(energys[0].get("energy")) if energys else None
            if energia_kwh is None:
                raise RuntimeError(f"Growatt no devolvió generación mensual válida para la planta {plant_id}")
            return energia_kwh
        except Exception as exc:  # noqa: BLE001
            ultimo_error = exc
            if intento < reintentos:
                time.sleep(min(1.5 * intento, 5.0))
    raise ultimo_error if ultimo_error is not None else RuntimeError("No se pudo obtener la generación mensual")


def obtener_metricas_mes_desde_smart_meter(
    api: growattServer.OpenApiV1,
    plant_id: int,
    fecha: date,
    fecha_fin: date,
    generacion_kwh: float | None,
    dataloggers: list[str] | None = None,
) -> dict[str, float | None]:
    if dataloggers is None:
        dataloggers = obtener_dataloggers_planta(api, plant_id)

    if not dataloggers:
        return {
            "generacion_kwh": generacion_kwh,
            "consumo_kwh": None,
            "exportacion_kwh": None,
        }

    muestras = []
    for datalogger_sn in dataloggers:
        payload = openapi_get_data(
            api,
            "device/ammeter/meter_list",
            {"datalog_sn": datalogger_sn},
            f"getting smart meter list for {datalogger_sn}",
        )
        meters = payload.get("meters", [])

        for meter in meters:
            address = str(meter.get("address") or "").strip()
            if not address:
                continue

            dia = fecha
            while dia <= fecha_fin:
                dia_str = dia.strftime("%Y-%m-%d")
                payload = openapi_get_data(
                    api,
                    "device/ammeter/meter_data",
                    {
                        "datalog_sn": datalogger_sn,
                        "address": address,
                        "start_date": dia_str,
                        "end_date": dia_str,
                    },
                    f"getting smart meter data for {datalogger_sn} address {address} on {dia_str}",
                )
                muestras.extend(payload.get("meter_data", []))
                dia += timedelta(days=1)
                time.sleep(0.15)

    importacion_red_kwh = resumir_historial_diario(muestras, "positiveActiveTodayEnergy")
    exportacion_kwh = resumir_historial_diario(muestras, "reverseActiveTodayEnergy")

    consumo_kwh = None
    if (
        generacion_kwh is not None
        and importacion_red_kwh is not None
        and exportacion_kwh is not None
    ):
        consumo_kwh = round(generacion_kwh + importacion_red_kwh - exportacion_kwh, 2)

    return {
        "generacion_kwh": generacion_kwh,
        "consumo_kwh": consumo_kwh,
        "exportacion_kwh": exportacion_kwh,
    }


def obtener_dataloggers_planta(
    api: growattServer.OpenApiV1,
    plant_id: int,
) -> list[str]:
    ultimo_error = None
    dispositivos = None
    for intento in range(1, 6):
        try:
            dispositivos = api.device_list(plant_id).get("devices", [])
            ultimo_error = None
            break
        except Exception as exc:  # noqa: BLE001
            ultimo_error = exc
            if intento < 5:
                time.sleep(min(1.5 * intento, 5.0))
    if dispositivos is None:
        raise ultimo_error if ultimo_error is not None else RuntimeError("No se pudo obtener la lista de dispositivos")

    dataloggers = []
    for dispositivo in dispositivos:
        if dispositivo.get("type") != 3:
            continue
        datalogger_sn = str(dispositivo.get("datalogger_sn") or "").strip()
        if datalogger_sn and datalogger_sn not in dataloggers:
            dataloggers.append(datalogger_sn)

    return dataloggers


def describir_metricas_faltantes_smart_meter(
    api: growattServer.OpenApiV1,
    plant_id: int,
) -> str:
    try:
        dataloggers = obtener_dataloggers_planta(api, plant_id)
    except Exception as exc:  # noqa: BLE001
        return f"no se pudo leer device_list: {exc}"

    if not dataloggers:
        return "la planta no tiene smart meter/datalogger tipo 3 visible en Open API"

    return f"smart meter visible, pero sin muestras históricas útiles para esa planta/mes ({len(dataloggers)} datalogger/s)"


def obtener_metricas_mes_desde_dispositivo(
    api: growattServer.OpenApiV1,
    plant_id: int,
    fecha: date,
    fecha_fin: date,
) -> dict[str, float | None]:
    return {
        "generacion_kwh": None,
        "consumo_kwh": None,
        "exportacion_kwh": None,
    }


def obtener_todas_las_plantas_openapi(api: growattServer.OpenApiV1, perpage: int = 100) -> list[dict]:
    plantas = []
    pagina = 1
    total_esperado = None

    while True:
        data = openapi_get_data(
            api,
            "plant/list",
            {
                "page": pagina,
                "perpage": perpage,
                "search_type": "",
                "search_keyword": "",
            },
            f"getting plant list page {pagina}",
        )
        pagina_plantas = data.get("plants", [])

        if total_esperado is None:
            total_esperado = int(data.get("count", len(pagina_plantas)))

        if not pagina_plantas:
            break

        plantas.extend(pagina_plantas)
        if len(plantas) >= total_esperado:
            break

        pagina += 1

    return plantas


def cargar_allowlist(path: Path) -> set[int] | None:
    if not path.exists():
        return None

    ids = set()
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        if "plant_id" not in (reader.fieldnames or []):
            raise ValueError(f"El archivo {path} debe tener una columna 'plant_id'.")
        for fila in reader:
            valor = str(fila.get("plant_id", "")).strip()
            if not valor:
                continue
            ids.add(int(valor))
    return ids if ids else set()


def guardar_inventario_plantas(path: Path, plantas: list[dict]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(["plant_id", "name", "status", "peak_power", "city", "country"])
        for planta in plantas:
            writer.writerow([
                planta.get("plant_id"),
                planta.get("name"),
                planta.get("status"),
                planta.get("peak_power"),
                planta.get("city"),
                planta.get("country"),
            ])
    return path


def extraer_energia_mes_token(config: Config, fecha: date) -> list[dict]:
    print(f"\nConectando a Growatt Open API para {fecha.strftime('%Y-%m')}...")
    api = growattServer.OpenApiV1(config.api_token)
    plantas = obtener_todas_las_plantas_openapi(api)
    inventario_path = guardar_inventario_plantas(Path("data/growatt_plantas_detectadas.csv"), plantas)

    if not plantas:
        print("No se encontraron plantas en la cuenta.")
        return []

    print(f"Plantas visibles para el token: {len(plantas)}")
    print(f"Inventario guardado en: {inventario_path}")

    if config.listar_plantas:
        for planta in plantas:
            print(f"{planta.get('plant_id')} | {planta.get('name')}")
        return []

    allowlist_ids = cargar_allowlist(config.allowlist_path)
    if allowlist_ids is None:
        raise RuntimeError(
            "No existe la allowlist de plantas autorizadas. "
            f"Creá {config.allowlist_path} con al menos una columna 'plant_id'."
        )
    if not allowlist_ids:
        raise RuntimeError(
            f"La allowlist {config.allowlist_path} está vacía. "
            "Agregá una columna 'plant_id' con una fila por cada planta autorizada."
        )

    plantas = [planta for planta in plantas if int(planta.get("plant_id")) in allowlist_ids]
    print(f"Plantas autorizadas por allowlist: {len(plantas)}")

    if not plantas:
        print("La allowlist no coincide con ninguna planta visible.")
        return []

    filas = []
    mes = fecha.strftime("%Y-%m")
    fecha_fin = ultimo_dia_mes(fecha)

    total_plantas = len(plantas)
    for idx, planta in enumerate(plantas, start=1):
        plant_id = planta.get("plant_id")
        plant_name = planta.get("name") or "Sin nombre"
        progreso = formatear_progreso(idx, total_plantas)

        try:
            energia_hist = None
            metricas_dispositivo = {
                "generacion_kwh": None,
                "consumo_kwh": None,
                "exportacion_kwh": None,
            }
            aviso_metricas = None
            ultimo_error = None
            for intento in range(1, config.reintentos + 1):
                try:
                    energia_kwh = obtener_generacion_mensual_directa(
                        config.api_token,
                        plant_id,
                        fecha,
                        fecha_fin,
                        reintentos=max(config.reintentos, 5),
                    )
                    metricas_dispositivo = obtener_metricas_mes_desde_smart_meter(
                        api,
                        plant_id,
                        fecha,
                        fecha_fin,
                        energia_kwh,
                    )
                    energia_hist = {"energia_kwh": energia_kwh}
                    ultimo_error = None
                    break
                except Exception as exc:  # noqa: BLE001
                    ultimo_error = exc
                    if intento < config.reintentos:
                        time.sleep(max(config.pausa_seg, 1.0) * intento)
            if energia_hist is None:
                raise ultimo_error if ultimo_error is not None else RuntimeError("No se pudo obtener energía")

            energia_kwh = metricas_dispositivo.get("generacion_kwh")

            fila = {
                "id_planta": plant_id,
                "cliente": str(plant_name).strip(),
                "mes": mes,
                "generacion_kwh": energia_kwh,
                "consumo_kwh": metricas_dispositivo.get("consumo_kwh"),
                "exportacion_kwh": metricas_dispositivo.get("exportacion_kwh"),
                "potencia_nominal_kw": a_float(planta.get("peak_power")),
                "estado": planta.get("status", "N/D"),
                "ultimo_dato": "",
            }
            filas.append(fila)
            generacion_legible = formatear_metrica(fila["generacion_kwh"])
            consumo_legible = formatear_metrica(fila["consumo_kwh"])
            exportacion_legible = formatear_metrica(fila["exportacion_kwh"])
            print(
                f"  {progreso} OK {plant_name:<35} "
                f"gen={generacion_legible} cons={consumo_legible} exp={exportacion_legible}"
            )
            if fila["consumo_kwh"] is None or fila["exportacion_kwh"] is None:
                aviso_metricas = describir_metricas_faltantes_smart_meter(api, plant_id)
                print(f"  {progreso} AVISO {plant_name}: {aviso_metricas}")
        except Exception as exc:  # noqa: BLE001
            print(f"  {progreso} ERROR {plant_name}: {exc}")
            filas.append(
                {
                    "id_planta": plant_id,
                    "cliente": str(plant_name).strip(),
                    "mes": mes,
                    "generacion_kwh": None,
                    "consumo_kwh": None,
                    "exportacion_kwh": None,
                    "potencia_nominal_kw": a_float(planta.get("peak_power")),
                    "estado": f"ERROR: {exc}",
                    "ultimo_dato": "",
                }
            )

        time.sleep(config.pausa_seg)

    return filas


def extraer_energia_mes(config: Config, fecha: date) -> list[dict]:
    if config.server_cookie:
        return extraer_energia_mes_cookie(config, fecha)
    if config.api_token:
        return extraer_energia_mes_token(config, fecha)

    print(f"\nConectando a Growatt para {fecha.strftime('%Y-%m')}...")
    api = growattServer.GrowattApi(add_random_user_id=True)

    resultado = api.login(config.usuario, config.password)
    if not resultado.get("userId"):
        raise RuntimeError("Error de login en Growatt. Revisá usuario y contraseña.")

    user_id = resultado["userId"]
    print(f"Login correcto. userId={user_id}")

    plantas = api.plant_list(user_id)
    if not plantas:
        print("No se encontraron plantas en la cuenta.")
        return []

    print(f"Plantas encontradas: {len(plantas)}")

    filas = []
    mes = fecha.strftime("%Y-%m")

    total_plantas = len(plantas)
    for idx, planta in enumerate(plantas, start=1):
        plant_id = planta.get("plantId") or planta.get("id")
        plant_name = planta.get("plantName") or planta.get("name") or "Sin nombre"
        progreso = formatear_progreso(idx, total_plantas)

        try:
            detalle = api.plant_detail(plant_id, timespan=growattServer.Timespan.month, date=fecha)

            energia_kwh = (
                detalle.get("valMap", {}).get("eMonth")
                or detalle.get("eMonth")
                or detalle.get("energy")
            )
            potencia_kw = detalle.get("nominalPower") or planta.get("nominalPower")
            estado = detalle.get("status", "N/D")
            ultimo_dato = detalle.get("lastUpdateTime", "N/D")

            fila = {
                "id_planta": plant_id,
                "cliente": str(plant_name).strip(),
                "mes": mes,
                "generacion_kwh": a_float(energia_kwh),
                "consumo_kwh": None,
                "exportacion_kwh": None,
                "potencia_nominal_kw": a_float(potencia_kw),
                "estado": estado,
                "ultimo_dato": ultimo_dato,
            }
            filas.append(fila)
            generacion_legible = formatear_metrica(fila["generacion_kwh"])
            consumo_legible = formatear_metrica(fila["consumo_kwh"])
            exportacion_legible = formatear_metrica(fila["exportacion_kwh"])
            print(
                f"  {progreso} OK {plant_name:<35} "
                f"gen={generacion_legible} cons={consumo_legible} exp={exportacion_legible}"
            )
        except Exception as exc:  # noqa: BLE001
            print(f"  {progreso} ERROR {plant_name}: {exc}")
            filas.append(
                {
                    "id_planta": plant_id,
                    "cliente": str(plant_name).strip(),
                    "mes": mes,
                    "generacion_kwh": None,
                    "consumo_kwh": None,
                    "exportacion_kwh": None,
                    "potencia_nominal_kw": a_float(planta.get("nominalPower")),
                    "estado": f"ERROR: {exc}",
                    "ultimo_dato": "N/D",
                }
            )

        time.sleep(config.pausa_seg)

    return filas


def extraer_energia_mes_cookie(config: Config, fecha: date) -> list[dict]:
    print(f"\nConectando a server.growatt.com con sesión web para {fecha.strftime('%Y-%m')}...")
    api = growattServer.GrowattApi(add_random_user_id=True)
    api.server_url = DEFAULT_SERVER_URL
    api.session.headers.update({"Cookie": config.server_cookie})

    try:
        response = api.session.post(
            api.get_url("newTwoPlantAPI.do"),
            params={"op": "getAllPlantListTwo"},
            data={
                "language": "1",
                "nominalPower": "",
                "order": "1",
                "pageSize": "15",
                "plantName": "",
                "plantStatus": "",
                "toPageNum": "1",
            },
            allow_redirects=False,
        )
        if config.diagnostico_cookie:
            print(f"HTTP status: {response.status_code}")
            print(f"Location: {response.headers.get('Location')}")
            print(response.text[:2000])
            return []
        data = response.json()
        plantas = data.get("PlantList", [])
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError(
            "No se pudo leer la sesión web de Growatt con la cookie provista. "
            "Probablemente expiró o faltan cookies."
        ) from exc

    if not plantas:
        raise RuntimeError(
            "La sesión web no devolvió plantas. Revisá que la cookie sea válida y pertenezca a una cuenta con acceso."
        )

    print(f"Plantas encontradas en sesión web: {len(plantas)}")
    filas = []
    mes = fecha.strftime("%Y-%m")

    total_plantas = len(plantas)
    for idx, planta in enumerate(plantas, start=1):
        plant_id = planta.get("plantId") or planta.get("id")
        plant_name = planta.get("plantName") or planta.get("plantNameStr") or planta.get("name") or "Sin nombre"
        progreso = formatear_progreso(idx, total_plantas)

        try:
            detalle = api.plant_detail(plant_id, timespan=growattServer.Timespan.month, date=fecha)
            energia_kwh = (
                detalle.get("valMap", {}).get("eMonth")
                or detalle.get("eMonth")
                or detalle.get("energy")
            )
            potencia_kw = detalle.get("nominalPower") or planta.get("nominalPower")
            estado = detalle.get("status", planta.get("plantStatus", "N/D"))
            ultimo_dato = detalle.get("lastUpdateTime", "N/D")

            fila = {
                "id_planta": plant_id,
                "cliente": str(plant_name).strip(),
                "mes": mes,
                "generacion_kwh": a_float(energia_kwh),
                "consumo_kwh": None,
                "exportacion_kwh": None,
                "potencia_nominal_kw": a_float(potencia_kw),
                "estado": estado,
                "ultimo_dato": ultimo_dato,
            }
            filas.append(fila)
            generacion_legible = formatear_metrica(fila["generacion_kwh"])
            consumo_legible = formatear_metrica(fila["consumo_kwh"])
            exportacion_legible = formatear_metrica(fila["exportacion_kwh"])
            print(
                f"  {progreso} OK {plant_name:<35} "
                f"gen={generacion_legible} cons={consumo_legible} exp={exportacion_legible}"
            )
        except Exception as exc:  # noqa: BLE001
            print(f"  {progreso} ERROR {plant_name}: {exc}")
            filas.append(
                {
                    "id_planta": plant_id,
                    "cliente": str(plant_name).strip(),
                    "mes": mes,
                    "generacion_kwh": None,
                    "consumo_kwh": None,
                    "exportacion_kwh": None,
                    "potencia_nominal_kw": a_float(planta.get("nominalPower")),
                    "estado": f"ERROR: {exc}",
                    "ultimo_dato": "N/D",
                }
            )

        time.sleep(config.pausa_seg)

    return filas


def guardar_excel_auxiliar(filas: list[dict], fecha: date, destino: Path) -> Path:
    destino.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Growatt {fecha.strftime('%Y-%m')}"

    ws.merge_cells("A1:I1")
    ws["A1"] = f"Datos Growatt - {fecha.strftime('%Y-%m')}"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    columnas = [
        "id_planta",
        "cliente",
        "mes",
        "generacion_kwh",
        "consumo_kwh",
        "exportacion_kwh",
        "potencia_nominal_kw",
        "estado",
        "ultimo_dato",
    ]
    titulos = {
        "id_planta": "ID Planta",
        "cliente": "Cliente / Planta",
        "mes": "Mes",
        "generacion_kwh": "Generación (kWh)",
        "consumo_kwh": "Consumo (kWh)",
        "exportacion_kwh": "Exportación (kWh)",
        "potencia_nominal_kw": "Potencia nominal (kW)",
        "estado": "Estado",
        "ultimo_dato": "Último dato",
    }

    header_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, columna in enumerate(columnas, start=1):
        celda = ws.cell(row=2, column=col_idx, value=titulos[columna])
        celda.fill = header_fill
        celda.font = header_font
        celda.alignment = Alignment(horizontal="center")

    alt_fill = PatternFill("solid", fgColor="DEEAF1")
    for row_idx, fila in enumerate(filas, start=3):
        for col_idx, columna in enumerate(columnas, start=1):
            celda = ws.cell(row=row_idx, column=col_idx, value=fila.get(columna))
            celda.alignment = Alignment(horizontal="center")
            if row_idx % 2 == 1:
                celda.fill = alt_fill

    fila_total = len(filas) + 3
    ws.cell(row=fila_total, column=1, value="TOTAL").font = Font(bold=True)
    ws.merge_cells(f"A{fila_total}:C{fila_total}")

    total_generacion = round(
        sum(fila["generacion_kwh"] for fila in filas if fila.get("generacion_kwh") is not None),
        2,
    )
    total_celda = ws.cell(row=fila_total, column=4, value=total_generacion)
    total_celda.font = Font(bold=True)
    total_celda.fill = PatternFill("solid", fgColor="BDD7EE")

    anchos = [14, 35, 12, 18, 18, 18, 22, 18, 22]
    for idx, ancho in enumerate(anchos, start=1):
        letra = openpyxl.utils.get_column_letter(idx)
        ws.column_dimensions[letra].width = ancho

    ws_info = wb.create_sheet("Info")
    ws_info["A1"] = "Generado el"
    ws_info["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws_info["A2"] = "Mes consultado"
    ws_info["B2"] = fecha.strftime("%Y-%m")
    ws_info["A3"] = "Total plantas"
    ws_info["B3"] = len(filas)
    ws_info["A4"] = "Total generación (kWh)"
    ws_info["B4"] = total_generacion

    wb.save(destino)
    print(f"\nExcel auxiliar generado: {destino}")
    return destino


def cargar_excel_auxiliar(destino: Path) -> list[dict]:
    if not destino.exists():
        raise FileNotFoundError(f"No existe el Excel auxiliar: {destino}")

    # La fila 1 es un título fusionado; los encabezados reales empiezan en la fila 2.
    df = pd.read_excel(destino, sheet_name=0, header=1)
    df.columns = df.columns.str.strip().str.lower()
    columnas_renombradas = {
        "id planta": "id_planta",
        "cliente / planta": "cliente",
        "mes": "mes",
        "generación (kwh)": "generacion_kwh",
        "consumo (kwh)": "consumo_kwh",
        "exportación (kwh)": "exportacion_kwh",
        "potencia nominal (kw)": "potencia_nominal_kw",
        "estado": "estado",
        "último dato": "ultimo_dato",
    }
    df = df.rename(columns=columnas_renombradas)

    columnas_requeridas = ["cliente", "mes", "generacion_kwh", "consumo_kwh", "exportacion_kwh"]
    faltantes = [col for col in columnas_requeridas if col not in df.columns]
    if faltantes:
        raise ValueError(
            f'El Excel auxiliar "{destino}" no tiene las columnas necesarias: {", ".join(faltantes)}'
        )

    df = df[df["cliente"].notna()].copy()
    df["cliente"] = df["cliente"].astype(str).str.strip()
    df = df[(df["cliente"] != "") & (df["cliente"].str.upper() != "TOTAL")].copy()
    df["mes"] = df["mes"].astype(str).str.strip()

    filas = df.to_dict(orient="records")
    if not filas:
        raise ValueError(f'El Excel auxiliar "{destino}" no tiene filas válidas para importar.')
    return filas


def obtener_auxiliares_generados(base_dir: Path) -> list[Path]:
    patrones = [
        "growatt_clientes_*.xlsx",
        "growatt_historico_*.xlsx",
    ]
    candidatos: list[Path] = []
    for patron in patrones:
        candidatos.extend(base_dir.glob(patron))

    def clave_orden(path: Path) -> tuple[int, str]:
        nombre = path.name
        if nombre.startswith("growatt_clientes_"):
            match = re.search(r"(\d{4}-\d{2})", nombre)
            return (1, match.group(1) if match else nombre)
        if nombre.startswith("growatt_historico_"):
            match = re.search(r"_a_(\d{4}-\d{2})", nombre)
            return (2, match.group(1) if match else nombre)
        return (9, nombre)

    return sorted({path.resolve() for path in candidatos}, key=clave_orden)


def cargar_todos_los_auxiliares(base_dir: Path) -> tuple[list[dict], list[Path]]:
    archivos = obtener_auxiliares_generados(base_dir)
    if not archivos:
        raise FileNotFoundError(f"No se encontraron Excel auxiliares en {base_dir}")

    filas: list[dict] = []
    archivos_validos: list[Path] = []
    for archivo in archivos:
        try:
            filas.extend(cargar_excel_auxiliar(archivo))
            archivos_validos.append(archivo)
        except ValueError as exc:
            print(f"  AVISO se omite {archivo.name}: {exc}")

    if not archivos_validos:
        raise ValueError(f"No se encontraron Excel auxiliares compatibles en {base_dir}")

    return filas, archivos_validos


def actualizar_hoja_datos(xlsx_path: Path, filas_growatt: list[dict]) -> tuple[int, int]:
    if not xlsx_path.exists():
        raise FileNotFoundError(f"No existe el archivo base: {xlsx_path}")

    xls = pd.ExcelFile(xlsx_path)
    if "datos" not in xls.sheet_names:
        raise ValueError(f'El archivo "{xlsx_path}" no tiene la hoja "datos".')

    hojas_existentes = {}
    for sheet in xls.sheet_names:
        hojas_existentes[sheet] = pd.read_excel(
            xlsx_path,
            sheet_name=sheet,
            header=None if sheet == "constantes_globales" else 0,
        )

    df_datos = hojas_existentes["datos"].copy()
    df_datos.columns = df_datos.columns.str.strip().str.lower()

    columnas_requeridas = ["cliente", "mes", "generacion_kwh", "consumo_kwh", "exportacion_kwh"]
    faltantes = [col for col in columnas_requeridas if col not in df_datos.columns]
    if faltantes:
        raise ValueError(f"Faltan columnas en hoja datos: {', '.join(faltantes)}")

    for col in ["generacion_kwh", "consumo_kwh", "exportacion_kwh"]:
        df_datos[col] = pd.to_numeric(df_datos[col], errors="coerce").astype(float)

    df_nuevo = pd.DataFrame(filas_growatt)[["cliente", "mes", "generacion_kwh", "consumo_kwh", "exportacion_kwh"]].copy()
    df_nuevo = df_nuevo[df_nuevo["cliente"].notna()].copy()
    df_nuevo["cliente"] = df_nuevo["cliente"].astype(str).str.strip()
    df_nuevo = df_nuevo[df_nuevo["cliente"] != ""].copy()

    actualizados = 0
    insertados = 0

    indice_existente = {
        (str(fila["cliente"]).strip(), str(fila["mes"]).strip()): idx
        for idx, fila in df_datos.iterrows()
    }

    def resolver_valor(actual, nuevo):
        nuevo_num = a_float(nuevo)
        actual_num = a_float(actual)
        if nuevo_num is None:
            return actual_num
        return nuevo_num

    for _, fila in df_nuevo.iterrows():
        clave = (fila["cliente"], fila["mes"])
        if clave in indice_existente:
            idx = indice_existente[clave]
            df_datos.at[idx, "generacion_kwh"] = resolver_valor(
                df_datos.at[idx, "generacion_kwh"],
                fila["generacion_kwh"],
            )
            df_datos.at[idx, "consumo_kwh"] = resolver_valor(
                df_datos.at[idx, "consumo_kwh"],
                fila["consumo_kwh"],
            )
            df_datos.at[idx, "exportacion_kwh"] = resolver_valor(
                df_datos.at[idx, "exportacion_kwh"],
                fila["exportacion_kwh"],
            )
            actualizados += 1
        else:
            generacion = a_float(fila["generacion_kwh"])
            consumo = a_float(fila["consumo_kwh"])
            exportacion = a_float(fila["exportacion_kwh"])
            if generacion is None and consumo is None and exportacion is None:
                continue
            nueva_fila = {columna: pd.NA for columna in df_datos.columns}
            nueva_fila["cliente"] = fila["cliente"]
            nueva_fila["mes"] = fila["mes"]
            nueva_fila["generacion_kwh"] = generacion
            nueva_fila["consumo_kwh"] = consumo
            nueva_fila["exportacion_kwh"] = exportacion
            df_datos = pd.concat(
                [
                    df_datos,
                    pd.DataFrame([nueva_fila]),
                ],
                ignore_index=True,
            )
            insertados += 1

    df_datos = df_datos.sort_values(by=["mes", "cliente"], kind="stable").reset_index(drop=True)

    with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        for sheet in xls.sheet_names:
            if sheet == "datos":
                continue
            hojas_existentes[sheet].to_excel(
                writer,
                sheet_name=sheet,
                index=False,
                header=(sheet != "constantes_globales"),
            )
        df_datos.to_excel(writer, sheet_name="datos", index=False)

    return actualizados, insertados


def main() -> int:
    config = parse_args()
    fecha = obtener_fecha_inicio(config.mes)
    destino = config.salida or obtener_nombre_salida(fecha)

    if config.cargar_todos_auxiliares:
        filas, archivos = cargar_todos_los_auxiliares(config.xlsx_datos.parent)
        print("\nUsando Excel auxiliares existentes:")
        for archivo in archivos:
            print(f"  - {archivo}")
    elif config.solo_actualizar_datos:
        if not config.actualizar_datos:
            raise ValueError("Usá --actualizar-datos junto con --solo-actualizar-datos.")
        filas = cargar_excel_auxiliar(destino)
        print(f"\nUsando Excel auxiliar existente: {destino}")
    else:
        filas = extraer_energia_mes(config, fecha)
        if not filas:
            print("No hubo datos para exportar.")
            return 0
        guardar_excel_auxiliar(filas, fecha, destino)

    if config.actualizar_datos:
        actualizados, insertados = actualizar_hoja_datos(config.xlsx_datos, filas)
        print(
            f'Hoja "datos" actualizada en {config.xlsx_datos}: '
            f"{actualizados} registros actualizados, {insertados} registros nuevos."
        )
        alertas = detectar_alertas_exportacion_baja(filas)
        if alertas:
            print(f"Se detectaron {len(alertas)} cliente(s) con posible exportación cero configurada.")
            enviar_alerta_exportacion_baja(config, alertas)

    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        print("\nProceso cancelado por el usuario.")
        raise SystemExit(130)
    except Exception as exc:  # noqa: BLE001
        print(f"\nError: {exc}", file=sys.stderr)
        raise SystemExit(1)
